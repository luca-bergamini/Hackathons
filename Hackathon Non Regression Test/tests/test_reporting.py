"""Test per src.reporting — copertura di aggregate_scores, build_report, save_structured_output."""

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from src.reporting.main import (
    aggregate_scores,
    build_report,
    run,
    save_structured_output,
    upload_results_to_s3,
    _build_cost_map,
    _compute_operative_metrics,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _mock_eval_results():
    return {
        "per_record": [
            {
                "test_id": "t1", "model_id": "m-a", "task": "classification",
                "score": 0.9, "correct": True,
                "output_text": "pos", "legacy_output": "pos", "ground_truth": "pos",
            },
            {
                "test_id": "t2", "model_id": "m-a", "task": "classification",
                "score": 0.3, "correct": False,
                "output_text": "neg", "legacy_output": "pos", "ground_truth": "pos",
            },
            {
                "test_id": "t1", "model_id": "m-b", "task": "classification",
                "score": 0.8, "correct": True,
                "output_text": "pos", "legacy_output": "pos", "ground_truth": "pos",
            },
            {
                "test_id": "t3", "model_id": "m-a", "task": "context_qa",
                "score": 0.7, "correct": True,
                "output_text": "Rome", "legacy_output": "Roma", "ground_truth": "Rome",
            },
        ],
        "per_task_model": {},
        "per_model": {},
    }


def _mock_runner_results():
    return {
        "m-a": [
            {"test_id": "t1", "task": "classification", "success": True,
             "latency_ms": 150, "retries": 0, "input_tokens": 100, "output_tokens": 10},
            {"test_id": "t2", "task": "classification", "success": True,
             "latency_ms": 200, "retries": 1, "input_tokens": 120, "output_tokens": 15},
            {"test_id": "t3", "task": "context_qa", "success": True,
             "latency_ms": 300, "retries": 0, "input_tokens": 500, "output_tokens": 50},
        ],
        "m-b": [
            {"test_id": "t1", "task": "classification", "success": True,
             "latency_ms": 180, "retries": 0, "input_tokens": 100, "output_tokens": 12},
        ],
    }


def _mock_model_configs():
    return [
        {"model_id": "m-a", "cost_input_1m": 0.08, "cost_output_1m": 0.35},
        {"model_id": "m-b", "cost_input_1m": 0.18, "cost_output_1m": 0.70},
    ]


# ---------------------------------------------------------------------------
# aggregate_scores
# ---------------------------------------------------------------------------

class TestAggregateScores:
    def test_basic(self):
        agg = aggregate_scores(_mock_eval_results())
        assert "per_record" in agg
        assert "per_task_model" in agg
        assert "per_model" in agg
        assert "best_models" in agg

    def test_per_task_model(self):
        agg = aggregate_scores(_mock_eval_results())
        assert "classification|m-a" in agg["per_task_model"]
        info = agg["per_task_model"]["classification|m-a"]
        assert info["num_records"] == 2
        assert info["avg_score"] == 0.6  # (0.9 + 0.3) / 2

    def test_per_model(self):
        agg = aggregate_scores(_mock_eval_results())
        assert "m-a" in agg["per_model"]
        assert agg["per_model"]["m-a"]["num_records"] == 3

    def test_with_runner_and_costs(self):
        agg = aggregate_scores(
            _mock_eval_results(), _mock_runner_results(), _mock_model_configs()
        )
        assert "operative_metrics" in agg
        assert len(agg["operative_metrics"]) > 0

    def test_best_models(self):
        agg = aggregate_scores(_mock_eval_results())
        best = agg["best_models"]
        assert "per_task" in best
        assert "overall" in best
        assert best["overall"]["model_id"] != ""


# ---------------------------------------------------------------------------
# _build_cost_map
# ---------------------------------------------------------------------------

class TestBuildCostMap:
    def test_basic(self):
        cm = _build_cost_map(_mock_model_configs())
        assert "m-a" in cm
        assert cm["m-a"]["cost_input_1m"] == 0.08

    def test_none(self):
        assert _build_cost_map(None) == {}


# ---------------------------------------------------------------------------
# _compute_operative_metrics
# ---------------------------------------------------------------------------

class TestComputeOperativeMetrics:
    def test_basic(self):
        cost_map = _build_cost_map(_mock_model_configs())
        op = _compute_operative_metrics(_mock_runner_results(), cost_map)
        assert "m-a|classification" in op
        info = op["m-a|classification"]
        assert info["total_requests"] == 2
        assert info["num_retries"] == 1
        assert info["avg_latency_ms"] == 175.0


# ---------------------------------------------------------------------------
# build_report (Excel)
# ---------------------------------------------------------------------------

class TestBuildReport:
    def test_creates_excel(self, tmp_path):
        agg = aggregate_scores(
            _mock_eval_results(), _mock_runner_results(), _mock_model_configs()
        )
        out = str(tmp_path / "test.xlsx")
        build_report(agg, out)
        assert Path(out).exists()

    def test_excel_has_4_sheets(self, tmp_path):
        agg = aggregate_scores(
            _mock_eval_results(), _mock_runner_results(), _mock_model_configs()
        )
        out = str(tmp_path / "test.xlsx")
        build_report(agg, out)
        wb = openpyxl.load_workbook(out)
        assert "Overview" in wb.sheetnames
        assert "Best Models" in wb.sheetnames
        assert "Verdict" in wb.sheetnames
        assert "Metriche Operative" in wb.sheetnames

    def test_empty_data(self, tmp_path):
        agg = aggregate_scores({"per_record": [], "per_task_model": {}, "per_model": {}})
        out = str(tmp_path / "empty.xlsx")
        build_report(agg, out)
        assert Path(out).exists()


# ---------------------------------------------------------------------------
# save_structured_output (JSON)
# ---------------------------------------------------------------------------

class TestSaveStructuredOutput:
    def test_creates_json(self, tmp_path):
        agg = aggregate_scores(_mock_eval_results())
        out = str(tmp_path / "test.json")
        save_structured_output(agg, out)
        assert Path(out).exists()
        with open(out, encoding="utf-8") as f:
            data = json.load(f)
        assert "per_record" in data
        assert "per_model" in data


# ---------------------------------------------------------------------------
# upload_results_to_s3
# ---------------------------------------------------------------------------

class TestUploadResultsToS3:
    def test_skips_when_no_bucket(self, monkeypatch):
        monkeypatch.delenv("S3_BUCKET", raising=False)
        result = upload_results_to_s3(bucket="")
        assert result == {}

    def test_uploads_excel_and_json(self, tmp_path, monkeypatch):
        excel_path = str(tmp_path / "report.xlsx")
        json_path = str(tmp_path / "report.json")
        Path(excel_path).write_text("fake excel", encoding="utf-8")
        Path(json_path).write_text('{"test": true}', encoding="utf-8")

        mock_s3 = MagicMock()
        with patch("src.reporting.main.boto3.client", return_value=mock_s3):
            result = upload_results_to_s3(
                excel_path=excel_path,
                json_path=json_path,
                bucket="test-bucket",
                s3_prefix="results/test-run",
            )

        assert "excel" in result
        assert "json" in result
        assert mock_s3.upload_file.call_count == 2

    def test_uploads_insights(self, monkeypatch):
        mock_s3 = MagicMock()
        with patch("src.reporting.main.boto3.client", return_value=mock_s3):
            result = upload_results_to_s3(
                insight_results={"summary": "test insight"},
                bucket="test-bucket",
                s3_prefix="results/test-run",
            )

        assert "insights" in result
        mock_s3.put_object.assert_called_once()

    def test_handles_upload_failure(self, tmp_path, monkeypatch):
        excel_path = str(tmp_path / "report.xlsx")
        Path(excel_path).write_text("fake excel", encoding="utf-8")

        mock_s3 = MagicMock()
        mock_s3.upload_file.side_effect = Exception("Access denied")
        with patch("src.reporting.main.boto3.client", return_value=mock_s3):
            result = upload_results_to_s3(
                excel_path=excel_path,
                bucket="test-bucket",
            )

        assert "excel" not in result


# ---------------------------------------------------------------------------
# run() orchestrator
# ---------------------------------------------------------------------------

class TestRunOrchestrator:
    def test_run_creates_both_files(self, tmp_path):
        excel_path = str(tmp_path / "report.xlsx")
        json_path = str(tmp_path / "report.json")
        result_path = run(
            _mock_eval_results(),
            _mock_runner_results(),
            _mock_model_configs(),
            output_path=excel_path,
            json_output_path=json_path,
        )
        assert result_path == excel_path
        assert Path(excel_path).exists()
        assert Path(json_path).exists()

    def test_run_without_runner_results(self, tmp_path):
        excel_path = str(tmp_path / "report.xlsx")
        json_path = str(tmp_path / "report.json")
        run(
            _mock_eval_results(),
            output_path=excel_path,
            json_output_path=json_path,
        )
        assert Path(excel_path).exists()

    def test_run_with_s3_upload(self, tmp_path):
        excel_path = str(tmp_path / "report.xlsx")
        json_path = str(tmp_path / "report.json")
        mock_s3 = MagicMock()
        with patch("src.reporting.main.boto3.client", return_value=mock_s3), \
             patch.dict("os.environ", {"S3_BUCKET": "test-bucket"}):
            run(
                _mock_eval_results(),
                output_path=excel_path,
                json_output_path=json_path,
                upload_to_s3=True,
            )
        assert Path(excel_path).exists()
