"""Report Aggregation Agent — Step 5.

Produce:
- Output strutturato JSON con metriche per record, per task×modello, overall, operative.
- Report Excel con 4 sheet: Overview, Best Models, Verdict, Metriche Operative.
"""

import json
import logging
import os
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DEFAULT_OUTPUT_PATH = "deliverable_intermedio_2.xlsx"
DEFAULT_JSON_PATH = "report_results.json"
_UNKNOWN = "unknown"

# --- Duplicate string constants (SonarQube S1192) ---
_K_MODEL_ID = "model_id"
_K_OVERALL_SCORE = "overall_score"
_K_AVG_SCORE = "avg_score"
_K_PER_RECORD = "per_record"
_AWS_ACCOUNT_ID: str = os.environ.get("AWS_ACCOUNT_ID", "")

# Header styles
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)


def _style_header_row(ws, row: int, max_col: int):
    """Applica stile alle intestazioni."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws):
    """Regola larghezza colonne automaticamente."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ---------------------------------------------------------------------------
# 2.11 — aggregate_scores
# ---------------------------------------------------------------------------

def aggregate_scores(
    eval_results: dict,
    runner_results: dict | None = None,
    model_configs: list[dict] | None = None,
) -> dict:
    """Aggrega i risultati di valutazione in metriche strutturate.

    Args:
        eval_results: output di evaluation.run() con per_record, per_task_model, per_model.
        runner_results: output di runner.run(), dict model_id -> list[result].
        model_configs: lista config modelli con costi.

    Returns:
        dict con: per_record, per_task_model, per_model, operative_metrics, best_models.
    """
    per_record = eval_results.get(_K_PER_RECORD, [])

    # --- Metriche per (task, model) ---
    task_model_scores: dict[tuple[str, str], list] = defaultdict(list)
    for rec in per_record:
        key = (rec.get("task", _UNKNOWN), rec.get(_K_MODEL_ID, _UNKNOWN))
        task_model_scores[key].append(rec)

    per_task_model = {}
    for (task, model_id), records in task_model_scores.items():
        scores = [r.get("score", 0.0) for r in records]
        correct_count = sum(1 for r in records if r.get("correct"))
        avg_score = round(sum(scores) / len(scores), 3) if scores else 0.0
        accuracy = round(correct_count / len(records), 3) if records else 0.0
        per_task_model[f"{task}|{model_id}"] = {
            "task": task,
            _K_MODEL_ID: model_id,
            _K_AVG_SCORE: avg_score,
            "accuracy": accuracy,
            "num_records": len(records),
            "num_correct": correct_count,
        }

    # --- Metriche overall per modello ---
    model_scores: dict[str, list] = defaultdict(list)
    for rec in per_record:
        model_scores[rec.get(_K_MODEL_ID, _UNKNOWN)].append(rec)

    per_model = {}
    for model_id, records in model_scores.items():
        scores = [r.get("score", 0.0) for r in records]
        correct_count = sum(1 for r in records if r.get("correct"))
        avg_score = round(sum(scores) / len(scores), 3) if scores else 0.0
        per_model[model_id] = {
            _K_OVERALL_SCORE: avg_score,
            "num_records": len(records),
            "num_correct": correct_count,
            "accuracy": round(correct_count / len(records), 3) if records else 0.0,
        }

    # --- Metriche operative per (model, task) dal runner ---
    cost_map = _build_cost_map(model_configs)
    operative = _compute_operative_metrics(runner_results, cost_map) if runner_results else {}

    # --- Best models ---
    best_models = _compute_best_models(per_task_model, per_model)

    return {
        _K_PER_RECORD: per_record,
        "per_task_model": per_task_model,
        "per_model": per_model,
        "operative_metrics": operative,
        "best_models": best_models,
    }


def _build_cost_map(model_configs: list[dict] | None) -> dict:
    """Mappa model_id -> {cost_input_1m, cost_output_1m}."""
    if not model_configs:
        return {}
    return {
        m[_K_MODEL_ID]: {
            "cost_input_1m": m.get("cost_input_1m", 0.0),
            "cost_output_1m": m.get("cost_output_1m", 0.0),
        }
        for m in model_configs
        if _K_MODEL_ID in m
    }


def _compute_operative_metrics(runner_results: dict, cost_map: dict) -> dict:
    """Calcola metriche operative per (model, task) dai risultati del runner."""
    grouped: dict[tuple[str, str], list] = defaultdict(list)
    for model_id, results in runner_results.items():
        for r in results:
            task = r.get("task", _UNKNOWN)
            grouped[(model_id, task)].append(r)

    operative = {}
    for (model_id, task), records in grouped.items():
        total = len(records)
        successes = sum(1 for r in records if r.get("success"))
        errors = total - successes
        retries = sum(r.get("retries", 0) for r in records)
        latencies = [r.get("latency_ms", 0) for r in records if r.get("success")]
        avg_latency = round(sum(latencies) / len(latencies), 1) if latencies else 0.0
        total_input_tokens = sum(r.get("input_tokens", 0) for r in records)
        total_output_tokens = sum(r.get("output_tokens", 0) for r in records)

        costs = cost_map.get(model_id, {"cost_input_1m": 0.0, "cost_output_1m": 0.0})
        cost_input = total_input_tokens * costs["cost_input_1m"] / 1_000_000
        cost_output = total_output_tokens * costs["cost_output_1m"] / 1_000_000
        total_cost = round(cost_input + cost_output, 6)
        avg_cost = round(total_cost / total, 6) if total > 0 else 0.0

        operative[f"{model_id}|{task}"] = {
            "model": model_id,
            "task": task,
            "avg_latency_ms": avg_latency,
            "total_requests": total,
            "num_errors": errors,
            "num_retries": retries,
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_cost": total_cost,
            "avg_cost_per_request": avg_cost,
        }

    return operative


def _compute_best_models(per_task_model: dict, per_model: dict) -> dict:
    """Identifica il miglior modello per task e complessivo."""
    # Best per task
    task_best: dict[str, dict] = {}
    for _key, info in per_task_model.items():
        task = info["task"]
        if task not in task_best or info[_K_AVG_SCORE] > task_best[task][_K_AVG_SCORE]:
            task_best[task] = {
                _K_MODEL_ID: info[_K_MODEL_ID],
                _K_AVG_SCORE: info[_K_AVG_SCORE],
            }

    # Best overall
    best_overall = {_K_MODEL_ID: "", _K_OVERALL_SCORE: -1.0}
    for model_id, info in per_model.items():
        if info[_K_OVERALL_SCORE] > best_overall[_K_OVERALL_SCORE]:
            best_overall = {_K_MODEL_ID: model_id, _K_OVERALL_SCORE: info[_K_OVERALL_SCORE]}

    return {"per_task": task_best, "overall": best_overall}


# ---------------------------------------------------------------------------
# 2.12 — build_report (Excel 4 sheet)
# ---------------------------------------------------------------------------

def build_report(aggregated: dict, output_path: str = DEFAULT_OUTPUT_PATH) -> str:
    """Genera il report Excel con 4 sheet obbligatori."""
    wb = openpyxl.Workbook()

    _build_overview_sheet(wb, aggregated)
    _build_best_models_sheet(wb, aggregated)
    _build_verdict_sheet(wb, aggregated)
    _build_operative_sheet(wb, aggregated)

    # Rimuovi il foglio vuoto di default se necessario
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(output_path)
    logger.info("Report Excel salvato in %s", output_path)
    return output_path


def _build_overview_sheet(wb: openpyxl.Workbook, agg: dict):
    """Sheet 1 — Overview: righe=modelli, colonne=task, celle=avg_score, colonna Overall."""
    ws = wb.create_sheet("Overview")

    per_task_model = agg.get("per_task_model", {})
    per_model = agg.get("per_model", {})

    # Raccogli tutti i task e modelli
    tasks = sorted({v["task"] for v in per_task_model.values()})
    models = sorted(per_model.keys())

    if not models:
        ws.append(["No data available"])
        return

    # Intestazione
    headers = ["Model"] + tasks + ["Overall"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    # Righe
    for model_id in models:
        row = [model_id]
        for task in tasks:
            key = f"{task}|{model_id}"
            info = per_task_model.get(key, {})
            row.append(info.get(_K_AVG_SCORE, ""))
        row.append(per_model.get(model_id, {}).get(_K_OVERALL_SCORE, ""))
        ws.append(row)

    _auto_width(ws)


def _build_best_models_sheet(wb: openpyxl.Workbook, agg: dict):
    """Sheet 2 — Best Models: miglior modello per task + complessivo."""
    ws = wb.create_sheet("Best Models")

    best = agg.get("best_models", {})
    per_task = best.get("per_task", {})
    overall = best.get("overall", {})

    # Best per task
    ws.append(["Task", "Best Model", "Score"])
    _style_header_row(ws, 1, 3)

    for task in sorted(per_task.keys()):
        info = per_task[task]
        ws.append([task, info[_K_MODEL_ID], info[_K_AVG_SCORE]])

    # Riga vuota + best overall
    ws.append([])
    ws.append([
        "Overall Best",
        overall.get(_K_MODEL_ID, ""),
        overall.get(_K_OVERALL_SCORE, ""),
    ])
    row_num = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)

    _auto_width(ws)


def _compute_legacy_verdict(legacy_out: object, gt: object) -> str:
    """Compare legacy output with ground truth for CORRECT/INCORRECT verdict."""
    if gt is None or legacy_out is None:
        return "N/A"
    lo = str(legacy_out).strip().lower()
    go = str(gt).strip().lower()
    return "CORRECT" if lo == go or lo in go or go in lo else "INCORRECT"


def _build_by_test_dict(per_record: list[dict]) -> tuple[dict, list[str]]:
    """Build by_test mapping and all_models list from per_record."""
    all_models = sorted({r.get(_K_MODEL_ID, "") for r in per_record})
    by_test: dict[str, dict] = {}
    for rec in per_record:
        tid = rec.get("test_id", "")
        if tid not in by_test:
            by_test[tid] = {
                "task": rec.get("task", ""),
                "legacy_output": rec.get("legacy_output", ""),
                "ground_truth": rec.get("ground_truth", ""),
                "legacy_verdict": _compute_legacy_verdict(
                    rec.get("legacy_output"), rec.get("ground_truth"),
                ),
                "models": {},
            }
        model_id = rec.get(_K_MODEL_ID, "")
        by_test[tid]["models"][model_id] = {
            "output": str(rec.get("output_text", ""))[:500],
            "verdict": "CORRECT" if rec.get("correct") else "INCORRECT",
        }
    return by_test, all_models

def _build_verdict_sheet(wb: openpyxl.Workbook, agg: dict):
    """Sheet 3 — Verdict: riga per record con CORRECT/INCORRECT."""
    ws = wb.create_sheet("Verdict")

    per_record = agg.get(_K_PER_RECORD, [])
    if not per_record:
        ws.append(["No data available"])
        return

    by_test, all_models = _build_by_test_dict(per_record)

    headers = ["test_id", "task", "legacy_output", "legacy_verdict"]
    for m in all_models:
        headers.extend([f"{m} output", f"{m} verdict"])
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for tid in sorted(by_test.keys()):
        info = by_test[tid]
        row = [tid, info["task"], str(info["legacy_output"])[:500], info["legacy_verdict"]]
        for m in all_models:
            m_data = info["models"].get(m, {})
            row.extend([m_data.get("output", ""), m_data.get("verdict", "N/A")])
        ws.append(row)

    _auto_width(ws)


def _build_operative_sheet(wb: openpyxl.Workbook, agg: dict):
    """Sheet 4 — Metriche Operative: riga per (modello, task)."""
    ws = wb.create_sheet("Metriche Operative")

    operative = agg.get("operative_metrics", {})

    headers = [
        "Model", "Task", "Avg Latency (ms)", "Total Requests",
        "Errors", "Retries", "Input Tokens", "Output Tokens",
        "Total Cost ($)", "Avg Cost/Request ($)",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for _key, metrics in sorted(operative.items()):
        ws.append([
            metrics.get("model", ""),
            metrics.get("task", ""),
            metrics.get("avg_latency_ms", 0),
            metrics.get("total_requests", 0),
            metrics.get("num_errors", 0),
            metrics.get("num_retries", 0),
            metrics.get("input_tokens", 0),
            metrics.get("output_tokens", 0),
            metrics.get("total_cost", 0),
            metrics.get("avg_cost_per_request", 0),
        ])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# 2.13 — Output JSON strutturato
# ---------------------------------------------------------------------------

def save_structured_output(aggregated: dict, output_path: str = DEFAULT_JSON_PATH) -> str:
    """Salva l'output strutturato in JSON."""
    # Converti le chiavi tuple in stringhe se necessario
    serializable = json.loads(json.dumps(aggregated, default=str))
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)
    logger.info("Output strutturato JSON salvato in %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# 3.12 — S3 Results Storage
# ---------------------------------------------------------------------------

def _upload_excel_file(s3, bucket: str, s3_prefix: str, excel_path: str) -> str | None:
    """Upload Excel report; return S3 URI or None on failure."""
    key = f"{s3_prefix}/report.xlsx"
    try:
        _owner = {"ExpectedBucketOwner": _AWS_ACCOUNT_ID} if _AWS_ACCOUNT_ID else None
        s3.upload_file(excel_path, bucket, key, ExtraArgs=_owner)
        uri = f"s3://{bucket}/{key}"
        logger.info("Uploaded Excel to %s", uri)
        return uri
    except Exception as e:
        logger.warning("Failed to upload Excel to S3: %s", e)
        return None


def _upload_json_file(s3, bucket: str, s3_prefix: str, json_path: str) -> str | None:
    """Upload JSON report; return S3 URI or None on failure."""
    key = f"{s3_prefix}/report_results.json"
    try:
        _owner = {"ExpectedBucketOwner": _AWS_ACCOUNT_ID} if _AWS_ACCOUNT_ID else None
        s3.upload_file(json_path, bucket, key, ExtraArgs=_owner)
        uri = f"s3://{bucket}/{key}"
        logger.info("Uploaded JSON to %s", uri)
        return uri
    except Exception as e:
        logger.warning("Failed to upload JSON to S3: %s", e)
        return None


def _upload_insights_object(s3, bucket: str, s3_prefix: str, insight_results: dict) -> str | None:
    """Upload insights dict as JSON; return S3 URI or None on failure."""
    key = f"{s3_prefix}/insights.json"
    try:
        body = json.dumps(insight_results, indent=2, default=str, ensure_ascii=False)
        _owner_kw = {"ExpectedBucketOwner": _AWS_ACCOUNT_ID} if _AWS_ACCOUNT_ID else {}
        s3.put_object(Bucket=bucket, Key=key, Body=body.encode("utf-8"), **_owner_kw)
        uri = f"s3://{bucket}/{key}"
        logger.info("Uploaded insights to %s", uri)
        return uri
    except Exception as e:
        logger.warning("Failed to upload insights to S3: %s", e)
        return None


def _resolve_s3_prefix(s3_prefix: str | None, job_id: str | None) -> str:
    """Return a deterministic S3 prefix: existing > job_id-based > timestamp-based."""
    if s3_prefix is not None:
        return s3_prefix
    if job_id:
        return f"results/{job_id}"
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"results/{ts}"


def upload_results_to_s3(
    excel_path: str | None = None,
    json_path: str | None = None,
    insight_results: dict | None = None,
    bucket: str | None = None,
    s3_prefix: str | None = None,
    job_id: str | None = None,
) -> dict:
    """Carica i risultati della pipeline su Amazon S3.

    BN_S3_PERSISTENCE: S3 keys are deterministic and idempotent.
    When job_id is provided, keys are always results/{job_id}/*, guaranteeing
    that re-running the same job overwrites rather than duplicates artefacts.

    Args:
        excel_path: percorso locale del report Excel.
        json_path: percorso locale del JSON strutturato.
        insight_results: dict output dell'Insight Agent (opzionale).
        bucket: nome del bucket S3 (default: da env S3_BUCKET).
        s3_prefix: prefisso S3 per i file (overrides job_id).
        job_id: deterministic job identifier for idempotent key generation.

    Returns:
        dict con le chiavi S3 dei file caricati.
    """
    if bucket is None:
        bucket = os.environ.get("S3_BUCKET", "")
    if not bucket:
        logger.warning("S3_BUCKET not set — skipping S3 upload")
        return {}

    region = os.environ.get("AWS_REGION", "eu-west-1")
    s3 = boto3.client("s3", region_name=region)
    uploaded: dict[str, str] = {}

    # BN_S3_PERSISTENCE: deterministic key from job_id; fallback to timestamp
    s3_prefix = _resolve_s3_prefix(s3_prefix, job_id)

    if excel_path and Path(excel_path).exists():
        uri = _upload_excel_file(s3, bucket, s3_prefix, excel_path)
        if uri:
            uploaded["excel"] = uri

    if json_path and Path(json_path).exists():
        uri = _upload_json_file(s3, bucket, s3_prefix, json_path)
        if uri:
            uploaded["json"] = uri

    if insight_results:
        uri = _upload_insights_object(s3, bucket, s3_prefix, insight_results)
        if uri:
            uploaded["insights"] = uri

    # BN_S3_PERSISTENCE: upload manifest.json listing all artefacts
    if uploaded:
        _upload_manifest(s3, bucket, s3_prefix, uploaded, job_id)

    logger.info("S3 upload completed: %d files uploaded", len(uploaded))
    return uploaded


def _upload_manifest(s3, bucket: str, s3_prefix: str, uploaded: dict, job_id: str | None) -> None:
    """BN_S3_PERSISTENCE: Write manifest.json listing all uploaded artefacts.

    The manifest records S3 paths, sizes, and timestamps for each artefact,
    enabling later listing of past jobs by reading manifests from S3.
    """
    manifest_key = f"{s3_prefix}/manifest.json"
    artefacts = []
    for label, s3_uri in uploaded.items():
        # Try to get object size for the artefact
        key = s3_uri.replace(f"s3://{bucket}/", "")
        size = None
        try:
            head = s3.head_object(Bucket=bucket, Key=key)
            size = head.get("ContentLength")
        except Exception:
            pass
        artefacts.append({"label": label, "s3_uri": s3_uri, "size_bytes": size})

    manifest = {
        "job_id": job_id,
        "s3_prefix": s3_prefix,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "artefacts": artefacts,
    }
    try:
        body = json.dumps(manifest, indent=2, ensure_ascii=False)
        _owner_kw = {"ExpectedBucketOwner": _AWS_ACCOUNT_ID} if _AWS_ACCOUNT_ID else {}
        s3.put_object(Bucket=bucket, Key=manifest_key,
                      Body=body.encode("utf-8"), **_owner_kw)
        logger.info("Manifest written to s3://%s/%s", bucket, manifest_key)
    except Exception as e:
        logger.warning("Failed to upload manifest: %s", e)


def _read_manifest_safe(s3, bucket: str, key: str) -> dict | None:
    """Fetch and parse a manifest.json; return dict or None on any failure."""
    try:
        _owner_kw = {"ExpectedBucketOwner": _AWS_ACCOUNT_ID} if _AWS_ACCOUNT_ID else {}
        response = s3.get_object(Bucket=bucket, Key=key, **_owner_kw)
        return json.loads(response["Body"].read().decode("utf-8"))
    except Exception as e:
        logger.debug("Could not read manifest %s: %s", key, e)
        return None


def list_past_jobs(bucket: str | None = None) -> list[dict]:
    """BN_S3_PERSISTENCE: List past jobs by scanning manifests from S3.

    Reads all manifest.json files under results/*/manifest.json and returns
    a list of job summaries, suitable for populating the UI job list.
    """
    if bucket is None:
        bucket = os.environ.get("S3_BUCKET", "")
    if not bucket:
        logger.warning("S3_BUCKET not set — cannot list past jobs")
        return []

    region = os.environ.get("AWS_REGION", "eu-west-1")
    s3 = boto3.client("s3", region_name=region)
    jobs = []
    try:
        paginator = s3.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=bucket, Prefix="results/"):
            for obj in page.get("Contents", []):
                key = obj.get("Key", "")
                if key.endswith("/manifest.json"):
                    manifest = _read_manifest_safe(s3, bucket, key)
                    if manifest is not None:
                        jobs.append(manifest)
    except Exception as e:
        logger.warning("Failed to list past jobs from S3: %s", e)
    return sorted(jobs, key=lambda j: j.get("created_at", ""), reverse=True)


# ---------------------------------------------------------------------------
# Orchestratore
# ---------------------------------------------------------------------------

def run(
    eval_results: dict,
    runner_results: dict | None = None,
    model_configs: list[dict] | None = None,
    output_path: str = DEFAULT_OUTPUT_PATH,
    json_output_path: str = DEFAULT_JSON_PATH,
    upload_to_s3: bool | None = None,
    insight_results: dict | None = None,
) -> str:
    """Esegue l'intero Step 5: aggrega → report Excel → output JSON → (opt) S3."""
    aggregated = aggregate_scores(eval_results, runner_results, model_configs)

    # Report Excel
    build_report(aggregated, output_path)

    # Output JSON strutturato
    save_structured_output(aggregated, json_output_path)

    # S3 upload — auto-enable if S3_BUCKET is set
    should_upload = upload_to_s3 if upload_to_s3 is not None else bool(
        os.environ.get("S3_BUCKET")
    )
    if should_upload:
        upload_results_to_s3(
            excel_path=output_path,
            json_path=json_output_path,
            insight_results=insight_results,
        )

    return output_path


if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    parser = argparse.ArgumentParser(description="Reporting — Step 5")
    parser.add_argument("--eval-json", required=True)
    parser.add_argument("--runner-json", default=None)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--json-output", default=DEFAULT_JSON_PATH)
    args = parser.parse_args()

    with open(args.eval_json, encoding="utf-8") as f:
        eval_data = json.load(f)

    runner_data = None
    if args.runner_json:
        with open(args.runner_json, encoding="utf-8") as f:
            runner_data = json.load(f)

    run(eval_data, runner_data, output_path=args.output, json_output_path=args.json_output)
